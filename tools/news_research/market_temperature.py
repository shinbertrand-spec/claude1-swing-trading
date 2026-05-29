"""Market-temperature fetchers (Put-Call, CNN Fear & Greed, AAII, VIX term).

Per spec at ``Bertieboo/Output/2026-05-29-claude1-market-temperature-pass-implementation.md``:
overlay context for the news-research snapshot (schema 1.3). NEVER a gate —
``risk-and-compliance`` and the swing-critic panel use these as informational
input only, per [[ai-arbitrage-compression]].

Four fetchers (all fail-soft — never raise; return ``{"error": "<reason>",
"as_of": None}`` on failure) plus a top-level :func:`fetch_market_temperature`
composer. Per-fetcher TTL is enforced via a thin on-disk JSON cache at
``ledgers/news/_state/market_temperature_cache/<fetcher>.json``.

CLI::

    uv run python -m tools.news_research.market_temperature
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
import urllib.error
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable

TOOL = "tools/news_research/market_temperature.py"

DEFAULT_CACHE_DIR = Path("ledgers/news/_state/market_temperature_cache")

# Per-fetcher TTL in seconds. Values from spec § 3.1.
TTL_SECONDS: dict[str, int] = {
    "put_call": 3600,        # 1h
    "fear_greed": 3600,      # 1h
    "aaii": 86400,           # 24h (weekly cadence)
    "vix_term": 900,         # 15m
}

_DEFAULT_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

_HTTP_TIMEOUT = 10  # seconds


# ---------------------------------------------------------------------------
# Cache layer
# ---------------------------------------------------------------------------


def _now_ts() -> float:
    return time.time()


def _utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _cache_path(fetcher: str, *, cache_dir: Path) -> Path:
    return cache_dir / f"{fetcher}.json"


def cache_get(
    fetcher: str,
    *,
    cache_dir: Path = DEFAULT_CACHE_DIR,
    now_fn: Callable[[], float] = _now_ts,
) -> dict | None:
    """Return the cached value if present + fresh; None otherwise."""
    path = _cache_path(fetcher, cache_dir=cache_dir)
    if not path.is_file():
        return None
    try:
        wrapped = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    stored_at = float(wrapped.get("stored_at") or 0)
    ttl = int(wrapped.get("ttl_seconds") or 0)
    if now_fn() - stored_at > ttl:
        return None
    return wrapped.get("value")


def cache_set(
    fetcher: str,
    value: dict,
    *,
    cache_dir: Path = DEFAULT_CACHE_DIR,
    now_fn: Callable[[], float] = _now_ts,
) -> None:
    """Write the value to the cache with the fetcher's TTL.

    Never raises — cache failures must not break a successful fetch.
    """
    try:
        cache_dir.mkdir(parents=True, exist_ok=True)
        path = _cache_path(fetcher, cache_dir=cache_dir)
        path.write_text(
            json.dumps({
                "stored_at": now_fn(),
                "ttl_seconds": TTL_SECONDS.get(fetcher, 0),
                "value": value,
            }),
            encoding="utf-8",
        )
    except OSError:
        return


# ---------------------------------------------------------------------------
# HTTP helper (urllib — matches project convention from screener.py +
# thematic_portfolio/corpus/press_rss.py; no new deps)
# ---------------------------------------------------------------------------


def _http_get(
    url: str,
    *,
    headers: dict[str, str] | None = None,
    timeout: int = _HTTP_TIMEOUT,
) -> bytes:
    """GET ``url`` and return the raw body. Raises on non-2xx / network error."""
    base_headers = {"User-Agent": _DEFAULT_UA, "Accept": "*/*"}
    if headers:
        base_headers.update(headers)
    req = urllib.request.Request(url, headers=base_headers)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read()


# ---------------------------------------------------------------------------
# Error sentinel
# ---------------------------------------------------------------------------


def _err(reason: str) -> dict:
    return {"error": reason, "as_of": None}


# ---------------------------------------------------------------------------
# Fetcher 1 — Fear & Greed (CNN)
# ---------------------------------------------------------------------------


_FEAR_GREED_URL = "https://production.dataviz.cnn.io/index/fearandgreed/graphdata"

# CNN's dataviz endpoint enforces an Origin/Referer CORS-flavoured check
# server-side. Plain UA gets a 403; matching the CNN F&G page passes.
_CNN_HEADERS = {
    "Origin": "https://www.cnn.com",
    "Referer": "https://www.cnn.com/markets/fear-and-greed",
    "Accept": "application/json, text/plain, */*",
}


def _classify_fear_greed(value: int) -> str:
    if value <= 24:
        return "extreme_fear"
    if value <= 44:
        return "fear"
    if value <= 55:
        return "neutral"
    if value <= 75:
        return "greed"
    return "extreme_greed"


def fetch_fear_greed(
    *,
    http_get: Callable[..., bytes] = _http_get,
) -> dict:
    """CNN Fear & Greed composite. Daily cadence (cached 1h).

    v1.1: passes Origin + Referer matching cnn.com (the dataviz endpoint
    server-side-enforces CORS-flavoured origin checks; bare UA 403s).
    """
    try:
        raw = http_get(_FEAR_GREED_URL, headers=_CNN_HEADERS)
        doc = json.loads(raw)
        fg = doc.get("fear_and_greed") or {}
        score = fg.get("score")
        if score is None:
            return _err("missing_score_field")
        value = int(round(float(score)))
        ts = fg.get("timestamp")
        # CNN returns ISO; pass through if string, else compose from now.
        as_of = ts if isinstance(ts, str) else _utc_iso()
        return {
            "as_of": as_of,
            "value": value,
            "regime": _classify_fear_greed(value),
            "source": "cnn",
        }
    except (urllib.error.URLError, OSError, TimeoutError) as exc:
        return _err(f"network: {type(exc).__name__}")
    except (json.JSONDecodeError, ValueError, KeyError, TypeError) as exc:
        return _err(f"parse: {type(exc).__name__}")


# ---------------------------------------------------------------------------
# Fetcher 2 — Put-Call ratio (CBOE)
# ---------------------------------------------------------------------------


_PUT_CALL_URL_PRIMARY = (
    "https://cdn.cboe.com/api/global/us_indices/daily_prices/VIXPCC_History.csv"
)
# v1.1: fallback when the CDN CSV 403s. The daily market-statistics HTML page
# exposes the same numbers in a different shape; we scrape the most recent row.
_PUT_CALL_URL_FALLBACK = (
    "https://www.cboe.com/us/options/market_statistics/daily/"
)
_CBOE_HEADERS = {
    "Referer": "https://www.cboe.com/",
    "Origin": "https://www.cboe.com",
    "Accept": "text/csv, text/plain, text/html, */*",
}

# Fallback parser: the daily market-statistics page renders a table with
# rows like "TOTAL PUT/CALL RATIO 0.95" / "EQUITY PUT/CALL RATIO 0.62".
_PUT_CALL_FALLBACK_RE = re.compile(
    r"(TOTAL|EQUITY|INDEX)\s+PUT\s*/\s*CALL\s+RATIO[^0-9]{0,40}"
    r"([0-9]+(?:\.[0-9]+)?)",
    re.IGNORECASE,
)


def _parse_put_call_csv(text: str) -> dict:
    """Parse the CBOE VIXPCC_History.csv body."""
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if len(lines) < 2:
        return _err("empty_csv")
    header = [h.strip().lower() for h in lines[0].split(",")]
    last = lines[-1].split(",")
    if len(last) < 2:
        return _err("malformed_row")
    date_val = last[0]
    total: float | None = None
    equity: float | None = None
    for i, col in enumerate(header[1:], start=1):
        if i >= len(last):
            break
        try:
            val = float(last[i])
        except ValueError:
            continue
        if total is None and ("total" in col or i == 1):
            total = val
        if "equity" in col:
            equity = val
    if total is None:
        return _err("no_total_column")
    as_of = date_val
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            as_of = datetime.strptime(date_val, fmt).date().isoformat()
            break
        except ValueError:
            continue
    return {
        "as_of": as_of,
        "total": total,
        "equity": equity,
        "source": "cboe",
    }


def _parse_put_call_fallback_html(text: str) -> dict:
    """Scrape the CBOE daily market-statistics HTML for total + equity rows."""
    found: dict[str, float] = {}
    for m in _PUT_CALL_FALLBACK_RE.finditer(text):
        label = m.group(1).lower()
        try:
            found.setdefault(label, float(m.group(2)))
        except ValueError:
            continue
    total = found.get("total")
    if total is None:
        return _err("no_total_in_fallback_html")
    return {
        "as_of": datetime.now(timezone.utc).date().isoformat(),
        "total": total,
        "equity": found.get("equity"),
        "source": "cboe",
    }


def fetch_put_call_ratio(
    *,
    http_get: Callable[..., bytes] = _http_get,
) -> dict:
    """CBOE total + equity-only put/call ratios. Daily cadence (cached 1h).

    v1.1: try the primary CDN CSV with browser-like Referer + Origin first;
    on HTTPError or parse failure, fall through to the daily market-statistics
    HTML page and scrape the headline rows. Both paths fail-soft to ``_err()``.
    """
    primary_err: str | None = None
    try:
        raw = http_get(_PUT_CALL_URL_PRIMARY, headers=_CBOE_HEADERS)
        text = raw.decode("utf-8", errors="replace")
        parsed = _parse_put_call_csv(text)
        if "error" not in parsed:
            return parsed
        primary_err = f"primary_parse: {parsed['error']}"
    except (urllib.error.URLError, OSError, TimeoutError) as exc:
        primary_err = f"primary_network: {type(exc).__name__}"
    except (ValueError, IndexError) as exc:
        primary_err = f"primary_parse: {type(exc).__name__}"

    try:
        raw = http_get(_PUT_CALL_URL_FALLBACK, headers=_CBOE_HEADERS)
        text = raw.decode("utf-8", errors="replace")
        parsed = _parse_put_call_fallback_html(text)
        if "error" not in parsed:
            return parsed
        return _err(f"fallback_parse: {parsed['error']} (primary: {primary_err})")
    except (urllib.error.URLError, OSError, TimeoutError) as exc:
        return _err(f"fallback_network: {type(exc).__name__} (primary: {primary_err})")
    except (ValueError, IndexError) as exc:
        return _err(f"fallback_parse: {type(exc).__name__} (primary: {primary_err})")


# ---------------------------------------------------------------------------
# Fetcher 3 — AAII weekly survey
# ---------------------------------------------------------------------------


_AAII_XLS_URL = "https://www.aaii.com/files/surveys/sentiment.xls"
_AAII_HTML_URL = "https://www.aaii.com/sentimentsurvey/sent_results"

# Look for the percent lines in the HTML scrape. The page renders the
# current week's Bullish / Neutral / Bearish percentages in close proximity.
_AAII_PCT_RE = re.compile(
    r"(Bullish|Neutral|Bearish)\s*[^0-9%]{0,40}([0-9]{1,3}(?:\.[0-9]+)?)\s*%",
    re.IGNORECASE,
)

_AAII_DATE_RE = re.compile(
    r"Reported.*?(\d{1,2}/\d{1,2}/\d{2,4})", re.IGNORECASE | re.DOTALL
)


def _normalize_aaii_pct(val) -> float | None:
    """AAII files mix 0-1 floats and 0-100 percentages. Return 0-1 float."""
    try:
        f = float(val)
    except (TypeError, ValueError):
        return None
    if f > 1.0:
        f = f / 100.0
    if not (0 <= f <= 1):
        return None
    return f


def _aaii_rows_via_xlrd(raw: bytes) -> list[tuple] | None:
    """xlrd path for true legacy binary .xls (openpyxl can't open them)."""
    try:
        import xlrd  # type: ignore
    except ImportError:
        return None
    try:
        book = xlrd.open_workbook(file_contents=raw)
    except Exception:  # noqa: BLE001
        return None
    sheet = book.sheet_by_index(0)
    rows: list[tuple] = []
    for r in range(sheet.nrows):
        out_row: list = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            val = cell.value
            # xlrd cell type 3 = date stored as float; convert.
            if cell.ctype == 3:
                try:
                    tup = xlrd.xldate_as_tuple(val, book.datemode)
                    val = datetime(*tup) if any(tup[:3]) else None
                except Exception:  # noqa: BLE001
                    pass
            out_row.append(val)
        rows.append(tuple(out_row))
    return rows


def _parse_aaii_xls_bytes(raw: bytes) -> dict:
    """Parse the AAII sentiment.xls feed.

    Tries openpyxl (Excel 2007+ / .xlsx-internals) first; falls back to
    xlrd<2 for the true legacy binary .xls AAII currently publishes.
    """
    rows: list[tuple] | None = None
    opener_err: str | None = None
    try:
        from openpyxl import load_workbook
    except ImportError:
        opener_err = "openpyxl_unavailable"
    else:
        import io
        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                opener_err = "xls_no_sheet"
            else:
                rows = list(ws.iter_rows(values_only=True))
        except Exception as exc:  # noqa: BLE001 — openpyxl raises many shapes
            opener_err = f"xls_open: {type(exc).__name__}"
            rows = None

    # xlrd fallback for true legacy .xls (BadZipFile from openpyxl is the tell).
    if rows is None:
        xlrd_rows = _aaii_rows_via_xlrd(raw)
        if xlrd_rows is None:
            return _err(opener_err or "xls_open_failed")
        rows = xlrd_rows

    if not rows:
        return _err("xls_empty")

    header_row_idx: int | None = None
    cols: dict[str, int] = {}
    for i, row in enumerate(rows[:10]):
        labels = {
            (str(c).strip().lower() if c is not None else ""): j
            for j, c in enumerate(row)
        }
        date_keys = ("date", "reported date", "reported")
        if any(k in labels for k in date_keys) and (
            "bullish" in labels and "bearish" in labels
        ):
            header_row_idx = i
            for key in ("date", "reported date", "reported"):
                if key in labels:
                    cols["date"] = labels[key]
                    break
            cols["bullish"] = labels["bullish"]
            cols["bearish"] = labels["bearish"]
            if "neutral" in labels:
                cols["neutral"] = labels["neutral"]
            break

    if header_row_idx is None or "date" not in cols:
        return _err("xls_no_header")

    # Walk rows after the header and pick the LAST one that has all three
    # percentages set — files often pad with formulas / blank rows.
    bull = neutral = bear = None
    as_of_week: str | None = None
    for row in rows[header_row_idx + 1:]:
        b = _normalize_aaii_pct(row[cols["bullish"]]) if cols.get("bullish") is not None else None
        n = _normalize_aaii_pct(row[cols["neutral"]]) if cols.get("neutral") is not None else None
        x = _normalize_aaii_pct(row[cols["bearish"]]) if cols.get("bearish") is not None else None
        if b is None or x is None:
            continue
        bull, bear = b, x
        if n is not None:
            neutral = n
        date_val = row[cols["date"]]
        if isinstance(date_val, datetime):
            as_of_week = date_val.date().isoformat()
        elif date_val is not None:
            s = str(date_val).strip()
            for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
                try:
                    as_of_week = datetime.strptime(s, fmt).date().isoformat()
                    break
                except ValueError:
                    continue

    if bull is None or bear is None:
        return _err("xls_no_data_row")
    if neutral is None:
        neutral = round(max(0.0, 1.0 - bull - bear), 4)
    return {
        "as_of_week": as_of_week,
        "bull": bull,
        "neutral": neutral,
        "bear": bear,
        "bull_bear_spread": round(bull - bear, 4),
        "source": "aaii",
    }


def _fetch_aaii_xls(
    *, http_get: Callable[..., bytes] = _http_get,
) -> dict:
    """Primary path — canonical XLS. DOM-stable across page redesigns."""
    try:
        raw = http_get(_AAII_XLS_URL)
    except (urllib.error.URLError, OSError, TimeoutError) as exc:
        return _err(f"xls_network: {type(exc).__name__}")
    return _parse_aaii_xls_bytes(raw)


def _fetch_aaii_html(
    *, http_get: Callable[..., bytes] = _http_get,
) -> dict:
    """Fallback path — scrape sent_results HTML for Bullish/Neutral/Bearish %s."""
    try:
        raw = http_get(_AAII_HTML_URL)
        text = raw.decode("utf-8", errors="replace")
        found: dict[str, float] = {}
        for m in _AAII_PCT_RE.finditer(text):
            label = m.group(1).lower()
            pct = float(m.group(2)) / 100.0
            found.setdefault(label, pct)
        bull = found.get("bullish")
        neutral = found.get("neutral")
        bear = found.get("bearish")
        if bull is None or bear is None or neutral is None:
            return _err("missing_percentages")
        if not (0 <= bull <= 1 and 0 <= neutral <= 1 and 0 <= bear <= 1):
            return _err("percentages_out_of_range")
        as_of_week: str | None = None
        m = _AAII_DATE_RE.search(text)
        if m:
            datestr = m.group(1)
            for fmt in ("%m/%d/%Y", "%m/%d/%y"):
                try:
                    as_of_week = datetime.strptime(datestr, fmt).date().isoformat()
                    break
                except ValueError:
                    continue
        return {
            "as_of_week": as_of_week,
            "bull": bull,
            "neutral": neutral,
            "bear": bear,
            "bull_bear_spread": round(bull - bear, 4),
            "source": "aaii",
        }
    except (urllib.error.URLError, OSError, TimeoutError) as exc:
        return _err(f"network: {type(exc).__name__}")
    except (ValueError, AttributeError) as exc:
        return _err(f"parse: {type(exc).__name__}")


def fetch_aaii_survey(
    *,
    http_get: Callable[..., bytes] = _http_get,
) -> dict:
    """AAII weekly investor sentiment. Weekly cadence (cached 24h).

    v1.1: canonical XLS feed is the primary source (DOM-stable across page
    redesigns); HTML scrape stays as fallback when the XLS endpoint or
    parser fails. Both paths fail-soft to ``_err()``.
    """
    xls_result = _fetch_aaii_xls(http_get=http_get)
    if "error" not in xls_result:
        return xls_result
    html_result = _fetch_aaii_html(http_get=http_get)
    if "error" not in html_result:
        return html_result
    return _err(f"xls:{xls_result['error']};html:{html_result['error']}")


# ---------------------------------------------------------------------------
# Fetcher 4 — VIX term structure (CBOE CSVs)
# ---------------------------------------------------------------------------


_VIX_URL = (
    "https://cdn.cboe.com/api/global/us_indices/daily_prices/VIX_History.csv"
)
_VIX9D_URL = (
    "https://cdn.cboe.com/api/global/us_indices/daily_prices/VIX9D_History.csv"
)
_VIX3M_URL = (
    "https://cdn.cboe.com/api/global/us_indices/daily_prices/VIX3M_History.csv"
)


def _parse_cboe_last_close(text: str) -> tuple[str, float] | None:
    """Return (date_str, close) for the last CBOE history row."""
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if len(lines) < 2:
        return None
    header = [h.strip().lower() for h in lines[0].split(",")]
    last = lines[-1].split(",")
    if len(last) < 2 or len(last) != len(header):
        return None
    close_idx: int | None = None
    for i, col in enumerate(header):
        if col == "close":
            close_idx = i
            break
    if close_idx is None:
        # Fall back to last column.
        close_idx = len(last) - 1
    try:
        close = float(last[close_idx])
    except ValueError:
        return None
    date_val = last[0]
    for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            date_val = datetime.strptime(date_val, fmt).date().isoformat()
            break
        except ValueError:
            continue
    return date_val, close


def _classify_vix_term(vix: float, vix9d: float, vix3m: float) -> str:
    if vix3m <= 0 or vix <= 0:
        return "neutral"
    if vix > 0 and vix9d / vix > 1.10:
        return "short_term_stress"
    if vix / vix3m > 1.05:
        return "backwardation"
    if vix / vix3m < 0.95:
        return "contango"
    return "neutral"


def fetch_vix_term_structure(
    *,
    http_get: Callable[[str], bytes] = _http_get,
) -> dict:
    """VIX / VIX9D / VIX3M term-structure ratio. Cached 15m."""
    try:
        vix_raw = http_get(_VIX_URL).decode("utf-8", errors="replace")
        vix9d_raw = http_get(_VIX9D_URL).decode("utf-8", errors="replace")
        vix3m_raw = http_get(_VIX3M_URL).decode("utf-8", errors="replace")
        vix_pair = _parse_cboe_last_close(vix_raw)
        vix9d_pair = _parse_cboe_last_close(vix9d_raw)
        vix3m_pair = _parse_cboe_last_close(vix3m_raw)
        if not (vix_pair and vix9d_pair and vix3m_pair):
            return _err("malformed_csv")
        as_of = vix_pair[0]
        vix, vix9d, vix3m = vix_pair[1], vix9d_pair[1], vix3m_pair[1]
        if vix <= 0:
            return _err("non_positive_vix")
        return {
            "as_of": as_of,
            "vix": vix,
            "vix9d": vix9d,
            "vix3m": vix3m,
            "vix9d_vix": round(vix9d / vix, 4),
            "vix_vix3m": round(vix / vix3m, 4) if vix3m else None,
            "regime": _classify_vix_term(vix, vix9d, vix3m),
            "source": "cboe",
        }
    except (urllib.error.URLError, OSError, TimeoutError) as exc:
        return _err(f"network: {type(exc).__name__}")
    except (ValueError, IndexError) as exc:
        return _err(f"parse: {type(exc).__name__}")


# ---------------------------------------------------------------------------
# Cache wrappers
# ---------------------------------------------------------------------------


_FETCHERS: dict[str, Callable[..., dict]] = {
    "put_call": fetch_put_call_ratio,
    "fear_greed": fetch_fear_greed,
    "aaii": fetch_aaii_survey,
    "vix_term": fetch_vix_term_structure,
}


def _cached_fetch(
    name: str,
    *,
    http_get: Callable[[str], bytes] = _http_get,
    cache_dir: Path = DEFAULT_CACHE_DIR,
    now_fn: Callable[[], float] = _now_ts,
    use_cache: bool = True,
) -> dict:
    """Cache-wrap one fetcher. Cache hit on fresh, success-only values."""
    if use_cache:
        cached = cache_get(name, cache_dir=cache_dir, now_fn=now_fn)
        if cached is not None:
            return cached
    value = _FETCHERS[name](http_get=http_get)
    if "error" not in value:
        cache_set(name, value, cache_dir=cache_dir, now_fn=now_fn)
    return value


# ---------------------------------------------------------------------------
# Top-level composer
# ---------------------------------------------------------------------------


def fetch_market_temperature(
    *,
    http_get: Callable[[str], bytes] = _http_get,
    cache_dir: Path = DEFAULT_CACHE_DIR,
    now_fn: Callable[[], float] = _now_ts,
    use_cache: bool = True,
) -> dict:
    """Top-level composer. Calls all 4 fetchers, composes the snapshot block.

    Never raises. If all four fetchers fail, returns block with all-error
    children + populated ``errors[]``.
    """
    children: dict[str, dict] = {}
    errors: list[str] = []
    for name in ("put_call", "fear_greed", "aaii", "vix_term"):
        try:
            child = _cached_fetch(
                name,
                http_get=http_get,
                cache_dir=cache_dir,
                now_fn=now_fn,
                use_cache=use_cache,
            )
        except Exception as exc:  # defense-in-depth — fetchers already fail-soft
            child = _err(f"unexpected: {type(exc).__name__}: {exc}")
        children[name] = child
        if "error" in child:
            errors.append(f"{name}: {child['error']}")

    # Compose top-level as_of as the max of children's as_of (or as_of_week
    # for AAII), falling back to now when no child succeeded.
    candidates: list[str] = []
    for name, child in children.items():
        if "error" in child:
            continue
        v = child.get("as_of") or child.get("as_of_week")
        if isinstance(v, str):
            candidates.append(v)
    as_of = max(candidates) if candidates else _utc_iso()
    return {
        "as_of": as_of,
        "put_call": children["put_call"],
        "fear_greed": children["fear_greed"],
        "aaii": children["aaii"],
        "vix_term": children["vix_term"],
        "errors": errors,
    }


# ---------------------------------------------------------------------------
# Latest-snapshot loader (consumer-side helper)
# ---------------------------------------------------------------------------


DEFAULT_NEWS_ROOT = Path("ledgers/news")
STALE_SNAPSHOT_SECONDS = 2 * 3600  # 2h per spec § 3.4


def load_latest_market_temperature(
    *,
    news_root: Path = DEFAULT_NEWS_ROOT,
    now_fn: Callable[[], datetime] = lambda: datetime.now(timezone.utc),
) -> dict | None:
    """Return the freshest ``market_temperature`` block from disk, or None.

    Stale-snapshot detection per spec § 3.4: if the latest snapshot is more
    than 2 hours older than ``now_fn()`` OR all four fetchers in the block
    are in error state, return ``None`` so consumers receive a null block
    and notice the gap.
    """
    if not news_root.exists():
        return None
    try:
        import yaml
    except ImportError:
        return None
    day_dirs = sorted(
        (p for p in news_root.iterdir()
         if p.is_dir() and re.match(r"^\d{4}-\d{2}-\d{2}$", p.name)),
        reverse=True,
    )
    for day_dir in day_dirs:
        hh_files = sorted(
            (p for p in day_dir.iterdir()
             if p.is_file() and re.match(r"^\d{2}\.yml$", p.name)),
            reverse=True,
        )
        for hh_file in hh_files:
            try:
                doc = yaml.safe_load(hh_file.read_text(encoding="utf-8")) or {}
            except (OSError, yaml.YAMLError):
                continue
            block = doc.get("market_temperature")
            if not isinstance(block, dict):
                continue
            # All-errors → stale
            errors = block.get("errors") or []
            children = ("put_call", "fear_greed", "aaii", "vix_term")
            if len(errors) >= len(children):
                return None
            # Snapshot-age check via meta.asof / meta.fetched_at
            meta = doc.get("meta") or {}
            asof_str = meta.get("fetched_at") or meta.get("asof") or meta.get("snapshot_id")
            if isinstance(asof_str, str):
                try:
                    cleaned = asof_str.replace("Z", "+00:00")
                    asof = datetime.fromisoformat(cleaned)
                    if asof.tzinfo is None:
                        asof = asof.replace(tzinfo=timezone.utc)
                    if (now_fn() - asof).total_seconds() > STALE_SNAPSHOT_SECONDS:
                        return None
                except ValueError:
                    pass
            return block
    return None


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> int:
    p = argparse.ArgumentParser(
        prog="tools.news_research.market_temperature",
        description="Fetch the composed market-temperature block (4 fetchers).",
    )
    p.add_argument("--no-cache", action="store_true", help="Bypass the on-disk cache.")
    args = p.parse_args()
    block = fetch_market_temperature(use_cache=not args.no_cache)
    json.dump(block, sys.stdout, indent=2, default=str)
    sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
